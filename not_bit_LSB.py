import cv2
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image, ImageOps
from ssim import SSIM
from ssim.utils import get_gaussian_kernel
import openpyxl
import os
from tqdm import tqdm

#setting
gaussian_kernel_sigma = 1.5
gaussian_kernel_width = 11
gaussian_kernel_1d = get_gaussian_kernel(gaussian_kernel_width, gaussian_kernel_sigma)

#directory address  
folder_path = "C:\\Users\\Asus\\Documents\\Steganography_folder\\image"
file_excel_path = "C:\\Users\\Asus\\Documents\\Steganography_folder\\compare_image_max.xlsx"

#create file excel   
if os.path.isfile(file_excel_path):
    os.remove(file_excel_path)
wb = openpyxl.Workbook()
ws = wb.active
name = ['ori_image','MSE','PSNR','SSIM','CW-SSIM']
for i in range(5):
    ws.cell(row = 1, column = i + 1, value = name[i])
wb.save(file_excel_path)
#MSE
def MSE(imga, imgb):
    mse = np.mean((imga - imgb) ** 2)
    return mse

#PSNR
def PSNR(imga, imgb):
    if(MSE(imga, imgb) == 0):
        return 100
    psnr = cv2.PSNR(imga,imgb)
    return psnr

#compare
def compare_image(img_orig,img_new,file_excel_path,i):
    name_origi = os.path.basename(img_orig)
    name_new = os.path.basename(img_new)
    imga = cv2.imread(img_orig)
    imgb = cv2.imread(img_new)
    MSE_value = MSE(imga, imgb)
    PSNR_value = PSNR(imga, imgb)
    im1 = Image.open(img_orig)
    im2 = Image.open(img_new)
    img1 = ImageOps.grayscale(im1)
    plt.imshow(np.asarray(img1),cmap='gray',vmin=0,vmax=255)
    img2 = ImageOps.grayscale(im2)
    plt.imshow(np.asarray(img2),cmap='gray',vmin=0,vmax=255)
    SSIM_value = SSIM(img1, gaussian_kernel_1d).ssim_value(img2)
    CW_SSIM_value = SSIM(img1).cw_ssim_value(img2)
    wb = openpyxl.load_workbook(file_excel_path)
    sheet = wb['Sheet']
    for col in range(5):
        if col == 0:
            sheet.cell(row= i+1, column = 1, value = name_origi)
        if col == 1:
            sheet.cell(row= i+1, column = 2, value = round(MSE_value,5))
        if col == 2:
            sheet.cell(row= i+1, column = 3, value = round(PSNR_value,5))
        if col == 3:
            sheet.cell(row= i+1, column = 4, value = round(SSIM_value,5))
        sheet.cell(row= i+1, column = 5, value = round(CW_SSIM_value,5))
    wb.save(file_excel_path)

def not_bit(image_file):
    image = cv2.imread(image_file)
    (h_ima, w_ima, d_ima) = image.shape
    print (h_ima, w_ima)
    for h in range(h_ima):
        for w in range(w_ima):
            for d in range (d_ima):
                val = list(image[h,w])
                if (val[d]%2 == 0):
                    val[d] = val[d] + 1
                else:
                    val[d] = val[d] - 1
                image[h,w] = tuple(val)
    cv2.imwrite("abc.png",image)
num_image = 1
for file in tqdm(os.listdir(folder_path)):
    image_file= os.path.join(folder_path,file)
    not_bit(image_file)
    compare_image(image_file,'abc.png',file_excel_path,num_image)
    num_image += 1
    os.remove('abc.png')