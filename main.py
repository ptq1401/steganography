from LSBSteg import LSBSteg
import cv2
import os
import shutil
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image, ImageOps
from ssim import SSIM
from ssim.utils import get_gaussian_kernel
from tqdm import tqdm

#setting
gaussian_kernel_sigma = 1.5
gaussian_kernel_width = 11
gaussian_kernel_1d = get_gaussian_kernel(gaussian_kernel_width, gaussian_kernel_sigma)

#rename and move file
num_ima = 101
def rename_file(new_folder_path,change_str): 
    old_file_name= 'image.png'
    new_file_name= "stego_image_" + change_str + ".png" 
    global new_file_path
    new_file_path=os.path.join(new_folder_path,new_file_name)
    os.rename(old_file_name,new_file_path)
    global num_ima
    num_ima += 1

#directory address  
'''
new_folder_path = "C:\\Users\\Asus\\Documents\\folder_steg\\steg_image"
file_excel_path = "C:\\Users\\Asus\\Documents\\folder_steg\\compare_image.xlsx"
file_text_path = "C:\\Users\\Asus\\Documents\\folder_steg\\file.txt"
'''
def create(new_folder_path,file_excel_path):
    #create file excel   
    if os.path.isfile(file_excel_path):
        os.remove(file_excel_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    name = ['ori_imageiii','new_image','MSE','PSNR','SSIM','CW-SSIM']
    for i in range(6):
        ws.cell(row = 1, column = i + 1, value = name[i])
    wb.save(file_excel_path)

    #create folder
    if os.path.isdir(new_folder_path):
        shutil.rmtree(new_folder_path)
    os.mkdir(new_folder_path)

#encoding
def encode(new_folder_path,old_folder_path,file_excel_path,file_text_path):
    global notify
    notify = "Part of the message has been hidden"
    for file in tqdm(os.listdir(old_folder_path)):
        image_path= os.path.join(old_folder_path,file)
        steg = LSBSteg(cv2.imread(image_path))
        pointer = update_pointer()
        message = read_text_file(pointer,file_text_path)
        if message == "":
            notify = "All messages have been hidden"
            break
        img_encoded = steg.encode_text(message)
        cv2.imwrite("image.png", img_encoded)
        change_str=str(num_ima)
        rename_file(new_folder_path,change_str)
        compare_image(image_path,new_file_path,file_excel_path,num_ima-100)
    pointer = update_pointer()
    message = read_text_file(pointer,file_text_path)
    if message == "":
        notify = "All messages have been hidden"

'''
#decoding
def decode(new_folder_path):
    for file in os.listdir(new_folder_path):
        imagesteg_path= os.path.join(new_folder_path,file)
        im = cv2.imread(imagesteg_path)
        steg = LSBSteg(im)
        print("Text valueeee " + file +":",steg.decode_text())
'''

def update_notify():
    return notify

def read_text_file(i,file_text_path):
    f = open(file_text_path, 'r',encoding = 'utf-8')
    f.seek(i)
    message = ''
    while True:
        data = f.read(1)
        if data == '':
            global pointer
            pointer = f.tell()
            message += data
            break
        if (data == ' ')&((f.tell() - pointer)>1250):
            message += data
            pointer = f.tell()
            break
        message += data
    return message
pointer = 0
def update_pointer():
    return pointer
#MSE
def MSE(imga, imgb):
    mse = np.mean((imga - imgb) ** 2)
    return mse

#PSNR
def PSNR(imga, imgb):
    if(MSE(imga, imgb) == 0):
        return 100
    psnr = cv2.PSNR(imga,imgb)
    return psnr

#compare
def compare_image(img_orig,img_new,file_excel_path,i):
    name_origi = os.path.basename(img_orig)
    name_new = os.path.basename(img_new)
    imga = cv2.imread(img_orig)
    imgb = cv2.imread(img_new)
    MSE_value = MSE(imga, imgb)
    PSNR_value = PSNR(imga, imgb)
    im1 = Image.open(img_orig)
    im2 = Image.open(img_new)
    img1 = ImageOps.grayscale(im1)
    plt.imshow(np.asarray(img1),cmap='gray',vmin=0,vmax=255)
    img2 = ImageOps.grayscale(im2)
    plt.imshow(np.asarray(img2),cmap='gray',vmin=0,vmax=255)
    SSIM_value = SSIM(img1, gaussian_kernel_1d).ssim_value(img2)
    CW_SSIM_value = SSIM(img1).cw_ssim_value(img2)
    wb = openpyxl.load_workbook(file_excel_path)
    sheet = wb['Sheet']
    for col in range(6):
        if col == 0:
            sheet.cell(row= i, column = 1, value = name_origi)
        if col == 1:
            sheet.cell(row= i, column = 2, value = name_new)
        if col == 2:
            sheet.cell(row= i, column = 3, value = round(MSE_value,5))
        if col == 3:
            sheet.cell(row= i, column = 4, value = round(PSNR_value,5))
        if col == 4:
            sheet.cell(row= i, column = 5, value = round(SSIM_value,5))
        sheet.cell(row= i, column = 6, value = round(CW_SSIM_value,5))
    wb.save(file_excel_path)



